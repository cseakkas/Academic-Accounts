from django.shortcuts import render, redirect, get_object_or_404
from . import models 
# from django.contrib import messages
from accounts_app.decorators import UserLogin
from django.db.models import Q
import openpyxl, csv

from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from xhtml2pdf import pisa

# Create your views here.

year_array = ["2022", "2023", "2024", "2025", "2026", "2027", "2028", "2029", "2030"]
  
def check_user_permission(request, menu_url):
    chk_privilege    = models.UserAccessControl.objects.filter(user_id = int(request.session.get("user_id")), menu_id__menu_url = menu_url, menu_id__status = True, status = True).first()
    
    if chk_privilege: return chk_privilege
    else: None

def dashboardLogin(request):
    if request.session.get('user_id'):
        return redirect('/')
 
    if request.method == "POST":
        get_data = models.UserRegistration.objects.filter(email = request.POST['userEmail'], password = request.POST['userPassword']).first()
        if get_data:
            request.session['user_id'] = get_data.id
            request.session['user_email'] = get_data.email
            request.session['first_name'] = get_data.first_name
            request.session['user_name'] = get_data.full_name
            request.session['user_photo'] = str(get_data.photo)
            return redirect('/')
        else:
            return redirect('/adminlogin/') 

 
    return render(request, "dashboard/login.html")

def admin_logout(request):  
    request.session['user_id'] = False
    return redirect('/adminlogin/')


@UserLogin
def dashboardPage(request):
    # chk_permission   = check_user_permission(request,'/book-category-entry/')
    # if chk_permission and chk_permission.view_action and chk_permission.insert_action: 


    # else:
    #     return redirect('/accessDeny')

    return render(request, "dashboard/index.html")
@UserLogin
def accessDeny(request):   
    return render(request, 'dashboard/access_deny.html')

@UserLogin
def classAdd(request):
    if request.method == 'POST':
        class_name = request.POST.get('class_name')
        text_id = class_name.replace(" ", "").lower()
        short_name = request.POST.get('short_name')
        rank = request.POST.get('rank')

        chk_exist = models.ClassList.objects.filter(text_id=text_id).first()
        if not chk_exist:
            class_obj = models.ClassList(
                class_name=class_name, text_id=text_id,
                short_name=short_name, rank=rank
            )
            class_obj.save() 
            return redirect('/settings/class-list/')
        else: 
            messages = f'Class with this name already exists!'
            return render(request, 'dashboard/settings/class_add.html', {'messages': messages})
    return render(request, 'dashboard/settings/class_add.html')

@UserLogin
def classList(request): 
    classes = models.ClassList.objects.filter(status=True).order_by('rank')
     
    return render(request, 'dashboard/settings/class_list.html', {'classes': classes})
 

@UserLogin
def classDelete(request, text_id):   
    text_id = text_id
    print(text_id)
    classes = models.ClassList.objects.filter(text_id=text_id).first()
    if classes: 
        classes.delete() 
        return redirect('/settings/class-list/')
    else: 
        return redirect('/settings/class-list/')
         
 

@UserLogin
def studentAdd(request):  
    class_list = models.ClassList.objects.filter(status=True).order_by('rank')
    
    if request.method == 'POST':
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        roll = request.POST.get('roll')
        reg_no = request.POST.get('reg_no')
        date_of_birth = request.POST.get('date_of_birth')
        email = request.POST.get('email')
        phone_number = request.POST.get('phone_number')
        address = request.POST.get('address')
        class_name_id = request.POST.get('class_name')
        father_name = request.POST.get('father_name')
        mother_name = request.POST.get('mother_name')
        guardian_name = request.POST.get('guardian_name')

        # Save the new student
        student_obj = models.StudentList(
            first_name=first_name,
            last_name=last_name,
            roll=roll,
            reg_no=reg_no,
            date_of_birth=date_of_birth,
            email=email,
            phone_number=phone_number,
            address=address,
            class_name_id=class_name_id,
            father_name=father_name,
            mother_name=mother_name,
            guardian_name=guardian_name
        )
        student_obj.save() 
        return redirect('/settings/student-list/')

    return render(request, 'dashboard/settings/student_add.html', {'class_list': class_list})


@UserLogin
def studentList(request): 
    class_list = models.ClassList.objects.filter(status=True).order_by('rank')
    student_list = models.StudentList.objects.all().order_by('id') 
    if request.method == 'POST':
        search_name = request.POST.get('search_name').strip()
        class_id = request.POST.get('class_id')
        search_roll = request.POST.get('search_roll')
  
        # Filter the student list based on search criteria
        student_list = models.StudentList.objects.all()

        if search_name:
            student_list = student_list.filter(Q(first_name__icontains=search_name) | Q(last_name__icontains=search_name))

        if class_id:
            class_id = int(class_id)
            student_list = student_list.filter(class_name_id=class_id)

        if search_roll:
            student_list = student_list.filter(roll__icontains=search_roll)

        context = {
            'student_list': student_list,
            'class_list': class_list,
            'class_id': class_id,
            'search_name': search_name,
            'search_roll': search_roll,
        }
        return render(request, 'dashboard/settings/student_list.html',context)
    
    context = {
        'class_list': class_list,
        'student_list': student_list,

    }
    return render(request, 'dashboard/settings/student_list.html',context)


def export_students_xlsx(request):
    search_name = request.GET.get('search_name', '')
    search_roll = request.GET.get('search_roll', '')
    class_id = request.GET.get('class_id', '')
 
    # Filter the students based on search criteria
    students = models.StudentList.objects.all()

    if search_name:
        students = students.filter(first_name__icontains=search_name)
    if search_roll:
        students = students.filter(roll__icontains=search_roll)
    if class_id:
        class_id = int(class_id)
        students = students.filter(class_name_id=class_id) 
        print("class_", students)

    # Create an Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Students'

    # Define headers
    headers = ['First Name', 'Last Name', 'Class', 'Roll', 'Registration No', 'Email', 'Phone']
    ws.append(headers)

    # Add student data
    for student in students:
        ws.append([
            student.first_name, 
            student.last_name, 
            student.class_name.class_name, 
            student.roll, 
            student.reg_no, 
            student.email, 
            student.phone_number
        ])

    # Set column widths (optional)
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 20  # You can adjust the width as needed

    # Prepare response to return as Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=students.xlsx'

    # Save the workbook to the response
    wb.save(response)
    
    return response


@UserLogin
def studentUpdate(request, id):
    student = models.StudentList.objects.get(id=id)
    
    class_list = models.ClassList.objects.filter(status=True).order_by('rank')
    
    if request.method == 'POST':
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        roll = request.POST.get('roll')
        reg_no = request.POST.get('reg_no')
        date_of_birth = request.POST.get('date_of_birth')
        email = request.POST.get('email')
        phone_number = request.POST.get('phone_number')
        address = request.POST.get('address')
        class_name_id = request.POST.get('class_name')
        father_name = request.POST.get('father_name')
        mother_name = request.POST.get('mother_name')
        guardian_name = request.POST.get('guardian_name')
        
        # Update the student data
        student.first_name = first_name
        student.last_name = last_name
        student.roll = roll
        student.reg_no = reg_no
        student.date_of_birth = date_of_birth
        student.email = email
        student.phone_number = phone_number
        student.address = address
        student.class_name_id = class_name_id
        student.father_name = father_name
        student.mother_name = mother_name
        student.guardian_name = guardian_name
        student.save()
        
        return redirect('/settings/student-list/')
    
    context = {
       'student': student,
        'class_list': class_list,
    }
    return render(request, 'dashboard/settings/student_edit.html', context)


@UserLogin
def studentDetails(request, student_id):
    student = get_object_or_404(models.StudentList, student_id=student_id)
    
    context = {
        'student': student
    }
    return render(request, 'dashboard/settings/student_details.html', context)
 

from django.template.loader import render_to_string 
def print_student_profile(request, student_id):
    student = get_object_or_404(models.StudentList, student_id=student_id)
  
    # Render the HTML template for the PDF
    html_string = render_to_string('dashboard/settings/student_profile_pdf.html', {'student': student})
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="student_profile_{student_id}.pdf"'
    
    # Create PDF from the HTML
    pisa_status = pisa.CreatePDF(html_string, dest=response)
    
    # Check if PDF was created successfully
    if pisa_status.err:
        return HttpResponse('Error generating PDF', status=400)
    
    return response


@UserLogin
def ChartOfAccountsAdd(request): 
    head_list = models.FinancialHead.objects.filter(status=True).order_by('rank')
    
    if request.method == 'POST':
        head_types_id = request.POST.get('head_types')
        chart_name = request.POST.get('chart_name')
        description = request.POST.get('description')
        rank = request.POST.get('rank')

        # generate textId from chart_name
        text_id = chart_name.replace(" ", "").lower()

        # Save the new head 
        instantce = models.ChartOfAccounts(
            financial_head_id=head_types_id, text_id = text_id,
            chart_name=chart_name, description=description, rank=rank
        )
        instantce.save() 
        return redirect('/settings/chart-of-accounts-list/')

    context = {
        'head_list': head_list,
    }
    return render(request, 'dashboard/settings/chart_of_accounts_add.html', context)


@UserLogin
def ChartOfAccountsList(request): 
    data_list = models.ChartOfAccounts.objects.filter(status=True)

    context = {
        'data_list': data_list,
    }
    return render(request, 'dashboard/settings/chart_of_accounts_list.html', context)
 
@UserLogin
def StudentsFeesSetup(request):  
    class_list = models.ClassList.objects.filter(status=True).order_by('rank') 
    year_list = year_array 

    if request.method == 'POST':
        if 'search_fees' in request.POST:
            class_id = request.POST.get('class_id')
            year = request.POST.get('year')
            chk_exist = models.ClassWiseFeeSetup.objects.filter(class_name_id=class_id, year = year).exists()
            if chk_exist:
                messages = f'Fees setup for class {class_id} and year {year} already exists!'
                context = {
                    'messages': messages,
                    'class_list': class_list,
                    'year_list': year_list,
                    'class_id': int(class_id),
                    'year': year,
                }
                return render(request, 'dashboard/settings/students_fees_setup.html', context)
            else:
                data_list = models.ChartOfAccounts.objects.filter(status=True)
                context = {
                    'data_list': data_list,
                    'class_list': class_list,
                    'year_list': year_list,
                    'class_id': int(class_id),
                    'year': year,
                }
                return render(request, 'dashboard/settings/students_fees_setup.html', context)

        else:
            class_id = request.POST.get('class_id')
            year = request.POST.get('year')
            fees_type_ids = request.POST.getlist('fees_type_id')
            fess_amount  = request.POST.getlist('fess_amount')

            print("fees_type_ids:", fees_type_ids)
            print("fess_amount:", fess_amount)

            for i in range(len(fees_type_ids)):
                fees_obj = models.ClassWiseFeeSetup(
                    class_name_id=class_id, year=year,
                    fees_head_id=fees_type_ids[i], amount=fess_amount[i]
                )
                fees_obj.save()
            return redirect('/settings/students-fees-setup-list/')

    context = {
        'class_list': class_list, 
        'year_list': year_list, 
    }

    return render(request, 'dashboard/settings/students_fees_setup.html', context)


    
@UserLogin
def StudentsFeesSetupList(request): 
    
    class_list = models.ClassList.objects.filter(status=True).order_by('rank') 
    year_list = year_array 

    if request.method == 'POST':
        class_id = request.POST.get('class_id')
        year = request.POST.get('year')

        data_list = models.ClassWiseFeeSetup.objects.filter(class_name_id=class_id, year=year)
        context = {
            'data_list': data_list,
            'class_list': class_list,
            'year_list': year_list,
            'class_id': int(class_id),
            'year': year,
        }
        return render(request, 'dashboard/settings/students_fees_setup_list.html', context)
 
    
    context = { 
        'class_list': class_list,
        'year_list': year_list,
    }  
    return render(request, 'dashboard/settings/students_fees_setup_list.html', context)

    
@UserLogin
def AccessControlAdd(request): 
    user_id = request.session.get('user_id')
    
    user_list = models.UserRegistration.objects.all()

    if request.method == 'POST':
        user_name_id = request.POST.get('user_name_id')

        access_list = models.UserAccessControl.objects.filter(user_id=user_name_id)
        context = {
            'access_list':access_list,
            'user_list':user_list,
            'user_name_id':int(user_name_id),
        }
        return render(request, 'dashboard/settings/access_controll_add.html', context)
    else: 
        access_list = models.UserAccessControl.objects.filter(user_id=user_id)
        context = {
            'access_list':access_list,
            'user_list':user_list,
        }
        return render(request, 'dashboard/settings/access_controll_add.html', context)


    
    
@UserLogin
def AccessControlList(request): 
    user_id = request.session.get('user_id')
    
    user_list = models.UserRegistration.objects.all()

    if request.method == 'POST':
        user_name_id = request.POST.get('user_name_id')

        access_list = models.UserAccessControl.objects.filter(user_id=user_name_id)
        context = {
            'access_list':access_list,
            'user_list':user_list,
            'user_name_id':int(user_name_id),
        }
        return render(request, 'dashboard/settings/access_controll_list.html', context)
    else: 
        access_list = models.UserAccessControl.objects.filter(user_id=user_id)
        context = {
            'access_list':access_list,
            'user_list':user_list,
        }
        return render(request, 'dashboard/settings/access_controll_list.html', context)


@UserLogin
def usersAdd(request):
    if request.method == "POST":
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        full_name = f"{first_name} {last_name}"
        email = request.POST.get('email')
        mobile = request.POST.get('mobile')
        designation = request.POST.get('designation')
        photo = request.FILES.get('photo') 
        password = request.POST.get('password')

        # Create a new user
        new_user = models.UserRegistration.objects.create(
            first_name=first_name,last_name=last_name, full_name=full_name,
            email=email, mobile=mobile, designation=designation,
            photo=photo, username=email, password=password
        )
        new_user.save()
        # messages.success(request, 'User added successfully.')
        return redirect('/users/user-list/')
 
    return render(request, 'dashboard/users/user_add.html')


@UserLogin
def usersUpdate(request, id): 
    user = models.UserRegistration.objects.filter(id=id).first()
    if user:
        if request.method == "POST":
            user.first_name = request.POST.get('first_name')
            user.last_name = request.POST.get('last_name')
            user.full_name = f"{user.first_name} {user.last_name}"
            user.email = request.POST.get('email')
            user.mobile = request.POST.get('mobile')
            user.designation = request.POST.get('designation')
            user.username = request.POST.get('email')
            user.status = True if request.POST.get('status') else False

            if request.FILES.get('photo'):
                user.photo = request.FILES.get('photo')

            user.save()
            # messages.success(request, 'User updated successfully.')
            return redirect('/users/user-list/')

    return render(request, 'dashboard/users/user_edit.html', {'user': user})

@UserLogin
def usersList(request):
    users = models.UserRegistration.objects.all()

    return render(request, 'dashboard/users/user_list.html', {'users': users})

@UserLogin
def usersDelete(request, id):
    users = models.UserRegistration.objects.filter(id=id)
    if users:
        users.delete()
        # messages.success(request, 'User deleted successfully.')
        return redirect('/users/user-list/') 

 
@UserLogin
def studentsFeeCollection(request): 
    if request.method == 'POST':
        if "search_student" in request.POST: 
            year = request.POST.get('year')
            student_id = request.POST.get('student_id')
    
            chk_student = models.StudentList.objects.filter(student_id=student_id).first()
            if chk_student: 
                get_fee_types = models.ClassWiseFeeSetup.objects.filter(class_name_id=chk_student.class_name_id)
                if get_fee_types:
                    context = {
                        'year_list': year_array, 
                        'chk_student': chk_student, 
                        'get_fee_types': get_fee_types,
                        'year': year,
                        'student_id': student_id,
                    }  
                    return render(request, 'dashboard/students/fee_collection.html', context)
                else:
                    context = {
                        'year_list': year_array, 
                        'chk_student': chk_student, 
                        "error_msg": "Fee Types not found.",
                        'year': year,
                        'student_id': student_id,
                    }
                    return render(request, 'dashboard/students/fee_collection.html', context)
            else:
                context = {
                    'year_list': year_array, 
                    "error_msg": "Fee Types not found.",
                    'year': year,
                    'student_id': student_id,
                }
                return render(request, 'dashboard/students/fee_collection.html', context)
        else: 
            student_id = request.POST.get('student_id')
            year = request.POST.get('year')
            selected_fees = request.POST.getlist('selected_fees') 
            get_student = models.StudentList.objects.filter(student_id=student_id).first()
            if get_student:
                for fee_id in selected_fees:
                    paid_amount = request.POST.get(f'paid_amount_{fee_id}') 
                    models.StudentFeeCollection.objects.create(
                        student_name_id = get_student.id,
                        class_name_id = get_student.class_name_id,
                        fees_head_id = int(fee_id),
                        year = year,
                        fee_amount = paid_amount,
                        due_amount = 0,
                        discount_amount = 0 
                    ) 
                return redirect('/students/fee-collection/')
              
    else:
        context = { 
            'year_list': year_array, 
        } 
        return render(request, 'dashboard/students/fee_collection.html', context)



@UserLogin
def studentFeeCollectionList(request):
    student_fees = models.StudentFeeCollection.objects.all()

    context = {
        'student_fees':student_fees
    }
    return render(request, 'dashboard/students/fee_collection_list.html', context)
    